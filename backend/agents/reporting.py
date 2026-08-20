import io
import os
import json
import uuid
import zipfile
import tempfile
import subprocess
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from dotenv import load_dotenv

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), '.env'))
db_url = os.getenv("DATABASE_URL", "postgresql://user:password@localhost:5432/finsentinel")
engine = create_engine(db_url)
SessionLocal = sessionmaker(bind=engine)

def generate_executive_report(session_id: str) -> bytes:
    db = SessionLocal()
    from models import AuditLog, ModelTrustScore
    
    if session_id == "all":
        logs = db.query(AuditLog).order_by(AuditLog.seq_id.asc()).all()
    else:
        logs = db.query(AuditLog).filter(AuditLog.session_id == session_id).order_by(AuditLog.seq_id.asc()).all()
        
    investigations = [L for L in logs if L.record_type == "investigation"]
    variances = [L for L in logs if L.record_type == "variance_query"]
    
    trust_scores = db.query(ModelTrustScore).order_by(ModelTrustScore.timestamp.asc()).all()
    latest_score = trust_scores[-1] if trust_scores else None
    
    db.close()
    
    # --- 1. Prepare Data for PPTXGenJS ---
    pptx_data = {
        "kpis": {
            "cases_analyzed": len(investigations),
            "precision": f"{(latest_score.precision * 100):.1f}%" if latest_score else "N/A",
            "fpr": f"{(latest_score.false_positive_rate * 100):.1f}%" if latest_score else "N/A"
        },
        "flagged_cases": [],
        "actions": []
    }
    
    typology_counts = {}
    
    for inv in investigations:
        result = inv.result or {}
        payload = inv.payload or {}
        
        # Extract typology if available
        notes = result.get("investigation_notes", {})
        typology = notes.get("typology_match", "Unknown")
        
        # Consider it flagged if the AI recommended DENY/ESCALATE, or a human explicitly decided DENY/ESCALATE
        ai_action = result.get("recommended_action", "").upper()
        human_decision = str(result.get("human_decision", "")).upper()
        
        is_flagged = (
            "DENY" in ai_action or "ESCALATE" in ai_action or
            "DENY" in human_decision or "ESCALATE" in human_decision or
            "FLAG" in str(result).upper()
        )
        
        if is_flagged:
            typology_counts[typology] = typology_counts.get(typology, 0) + 1
            pptx_data["flagged_cases"].append({
                "case_id": inv.execution_id or str(uuid.uuid4())[:8],
                "account": payload.get("account_id", "Unknown"),
                "amount": payload.get("amount", 0.0),
                "risk_score": result.get("risk_score", "N/A"),
                "confidence": result.get("calibrated_confidence", "N/A"),
                "reason": human_decision if human_decision and human_decision != "NONE" else f"AI Recommended: {ai_action}"
            })
            
    # Add Typologies Data
    pptx_data["typologies"] = [{"name": k, "count": v} for k, v in typology_counts.items()]
    
    # Generate Dynamic Actions
    dynamic_actions = []
    if any("Account Takeover" in k or "ATO" in k for k in typology_counts.keys()):
        dynamic_actions.append("Mandate 2FA for all accounts exhibiting IP rotation velocity spikes.")
    if any("Structuring" in k for k in typology_counts.keys()):
        dynamic_actions.append("Audit all transactions near the $10k reporting threshold for potential smurfing.")
    if any("Unknown" in k for k in typology_counts.keys()) or not dynamic_actions:
        dynamic_actions.append("Review manual queue cases in the attached spreadsheet.")

    # Pull natural-language answers from variance query audit logs.
    # Use result["final_answer"] — NOT var.response, which on pre-flight
    # entries is None or a raw system prompt, causing SQL JSON to bleed
    # into the slide.
    for var in variances:
        answer = None
        if var.result and isinstance(var.result, dict):
            answer = var.result.get("final_answer")
        # Fallback: if old log stored the answer in response field and it's
        # not a JSON blob or empty, use it.
        if not answer and var.response:
            candidate = var.response.strip()
            if candidate and not candidate.startswith("{") and not candidate.startswith("[") and len(candidate) > 20:
                answer = candidate
        if answer:
            # Truncate very long LLM answers to a single clean sentence for the slide.
            first_sentence = answer.split(".")[0].strip()
            if len(first_sentence) > 10:
                dynamic_actions.append(first_sentence + ".")

    # Cap at 6 bullets so they fit inside the slide textbox without overflow.
    pptx_data["actions"] = dynamic_actions[:6]
    
    # Fetch Red Team Benchmark
    import glob
    reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "static", "reports", "redteam")
    redteam_files = glob.glob(os.path.join(reports_dir, "redteam_results_*.json"))
    if redteam_files:
        latest_file = max(redteam_files, key=os.path.getmtime)
        try:
            with open(latest_file, "r") as f:
                rt_data = json.load(f)
                
            rt_structuring = rt_data.get("scenarios", [])[0].get("variants", [])
            pptx_data["redteam"] = {
                "baseline_conf": rt_structuring[0].get("calibrated_confidence", 0.0),
                "evasion_conf": rt_structuring[-1].get("calibrated_confidence", 0.0)
            }
        except Exception:
            pptx_data["redteam"] = None
        
    # --- 2. Generate PPTX using Node.js ---
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json') as tmp_json:
        json.dump(pptx_data, tmp_json)
        json_path = tmp_json.name
        
    pptx_out_path = json_path.replace(".json", ".pptx")
    
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    script_path = os.path.join(script_dir, "scripts", "generate_pptx.js")
    
    subprocess.run(["node", script_path, json_path, pptx_out_path], check=True, cwd=script_dir)
    
    with open(pptx_out_path, "rb") as f:
        pptx_bytes = f.read()
        
    os.remove(json_path)
    os.remove(pptx_out_path)
    
    # --- 3. Build XLSX ---
    wb = openpyxl.Workbook()
    
    # Sheet 1: Case Details
    ws_cases = wb.active
    ws_cases.title = "Case Details"
    
    header = ["Case ID", "Account", "Amount", "Risk Score", "Calibrated Confidence", "Flagged Reason", "Decision Outcome", "Reviewed By"]
    ws_cases.append(header)
    
    for inv in investigations:
        result = inv.result or {}
        payload = inv.payload or {}
        
        row = [
            inv.execution_id or "N/A",
            payload.get("account_id", "Unknown"),
            payload.get("amount", 0.0),
            result.get("risk_score", "N/A"),
            result.get("calibrated_confidence", "N/A"),
            inv.response if "FLAG" in str(inv.result) else "Normal",
            result.get("recommended_action", inv.action_type),
            "System Automated"
        ]
        ws_cases.append(row)
        
    # Formatting
    for col in ws_cases.columns:
        ws_cases.column_dimensions[col[0].column_letter].width = 20
        
    tab = Table(displayName="CasesTable", ref=f"A1:H{max(2, ws_cases.max_row)}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showRowStripes=True)
    ws_cases.add_table(tab)
    
    # Sheet 2: Trust Score Trend
    ws_trust = wb.create_sheet(title="Trust Score Trend")
    ws_trust.append(["Timestamp", "Precision", "Recall", "FPR", "Sample Size"])
    
    for row_idx, ts in enumerate(trust_scores, start=2):
        cell_time = ws_trust.cell(row=row_idx, column=1, value=ts.timestamp)
        cell_time.number_format = 'YYYY-MM-DD HH:MM'
        
        cell_prec = ws_trust.cell(row=row_idx, column=2, value=ts.precision)
        cell_prec.number_format = '0.0%'
        
        cell_rec = ws_trust.cell(row=row_idx, column=3, value=ts.recall)
        cell_rec.number_format = '0.0%'
        
        cell_fpr = ws_trust.cell(row=row_idx, column=4, value=ts.false_positive_rate)
        cell_fpr.number_format = '0.0%'
        
        ws_trust.cell(row=row_idx, column=5, value=ts.sample_size)

    for col in ws_trust.columns:
        ws_trust.column_dimensions[col[0].column_letter].width = 15
        
    tab2 = Table(displayName="TrustTable", ref=f"A1:E{max(2, ws_trust.max_row)}")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_trust.add_table(tab2)
    
    # Add LineChart
    if len(trust_scores) > 1:
        chart = LineChart()
        chart.title = "Model Trust Score Trend"
        chart.y_axis.title = 'Percentage'
        chart.x_axis.title = 'Timestamp'
        
        data = Reference(ws_trust, min_col=2, min_row=1, max_col=4, max_row=ws_trust.max_row)
        cats = Reference(ws_trust, min_col=1, min_row=2, max_row=ws_trust.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws_trust.add_chart(chart, "G2")
        
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    
    # --- 4. Zip bundle ---
    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("investigation_details.xlsx", xlsx_io.getvalue())
        zf.writestr("executive_report.pptx", pptx_bytes)
        
    zip_io.seek(0)
    return zip_io.getvalue()
